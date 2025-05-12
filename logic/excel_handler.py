from typing import cast
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime, timedelta

import consts as c

from logic.logger import logger as lg


class ExcelHandler:
    """
    Класс для работы с Excel-файлом.

    Methods
    -------
    - check_file(start_col, end_col)
        Проверяет даты в указанном диапазоне столбцов и возвращает
        словарь с именами и просроченными датами.
    """

    def __init__(self) -> None:
        self.filepath: str = c.EMPTY_STRING

    def check_file(
        self,
        start_col: str,
        end_col: str
    ) -> dict[str, list[dict[str, str]]]:
        """
        Проверяет даты в указанном диапазоне столбцов и возвращает
        словарь с именами и просроченными датами.

        Parameters
        ----------
        - start_col: str
            Начальный столбец для проверки (например, "A").

        - end_col: str
            Конечный столбец для проверки (например, "C").

        Returns
        -------
        - result: dict
            Словарь, где ключами являются имена тех, для кого просрочены
            документы а значениями - списки словарей с типом документа и
            просроченными датами.
        """

        lg.info("Start checking the excel file...")
        lg.info(f"Trying to open the file '{self.filepath}'...")

        try:
            wb = load_workbook(self.filepath, data_only=True)
            ws = wb[c.EXCEL_SHEET_NAME]
            lg.info(f"File '{self.filepath}' opened successfully.")
        except FileNotFoundError:
            lg.error(f"File not found: '{self.filepath}'.")
            return {}
        except Exception as e:
            lg.error(f"An error occurred while opening the file: {e}.")
            return {}

        result = {}

        lg.info("Set the threshold and today's date.")

        today = datetime.today()
        lg.info(f"Today's date: {today}.")

        threshold = today + timedelta(days=c.TIME_TRESHOLD)
        lg.info(f"Threshold date: {threshold}.")

        name_col = c.EXCEL_NAME_COL
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)

        lg.info("Start checking the dates in the file.")
        for row in ws.iter_rows(
            min_row=c.EXCEL_START_ROW,
            min_col=start_idx,
            max_col=end_idx
        ):
            row_index = cast(int, row[0].row)

            # name - имя раотника, для которого проверяются даты
            name = str(ws.cell(row=row_index, column=name_col).value)
            if not name:
                continue

            row_result = []
            for idx, cell in enumerate(row, start=start_idx):
                if isinstance(cell.value, datetime):
                    if cell.value <= threshold:
                        lg.info(
                            "Found a date that is less than or equal to the "
                            "threshold."
                        )
                        # header - заголовок столбца, например "Scadenza". В
                        # заголовке указан тип документа, для которого
                        # проверяется дата
                        header = str(
                            ws.cell(row=c.HEADER_ROW, column=idx).value
                        )
                        lg.info(f"===Date expired for: {name}.")
                        lg.info(f"===Expired object is: {header}.")
                        row_result.append(
                            {header: cell.value.strftime("%Y-%m-%d")}
                        )
                        value = row_result[-1][header]
                        lg.info(
                            f"===Expired date is: {value}."
                        )

            if row_result:
                result[name] = row_result

        lg.info("Finished checking the file.")
        lg.info("Closing the file...")
        wb.close()

        return result
